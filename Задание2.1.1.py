from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00



# dict_salary = {2007: 38916, 2008: 43646, 2009: 42492, 2010: 43846, 2011: 47451, 2012: 48243, 2013: 51510,
#                             2014: 50658, 2015: 52696, 2016: 62675, 2017: 60935, 2018: 58335, 2019: 69467, 2020: 73431,
#                             2021: 82690, 2022: 91795}
# dict_count = {2007: 2196, 2008: 17549, 2009: 17709, 2010: 29093, 2011: 36700, 2012: 44153, 2013: 59954,
#                            2014: 66837, 2015: 70039, 2016: 75145, 2017: 82823, 2018: 131701, 2019: 115086, 2020: 102243,
#                            2021: 57623, 2022: 18294}
# dict_salary_vac = {2007: 43770, 2008: 50412, 2009: 46699, 2010: 50570, 2011: 55770, 2012: 57960,
#                                 2013: 58804, 2014: 62384, 2015: 62322, 2016: 66817, 2017: 72460, 2018: 76879,
#                                 2019: 85300, 2020: 89791, 2021: 100987, 2022: 116651}
# dict_count_vac = {2007: 317, 2008: 2460, 2009: 2066, 2010: 3614, 2011: 4422, 2012: 4966, 2013: 5990,
#                                2014: 5492, 2015: 5375, 2016: 7219, 2017: 8105, 2018: 10062, 2019: 9016, 2020: 7113,
#                                2021: 3466, 2022: 1115}
# dict_salary_area = {'Москва': 76970, 'Санкт-Петербург': 65286, 'Новосибирск': 62254, 'Екатеринбург': 60962,
#                            'Казань': 52580, 'Краснодар': 51644, 'Челябинск': 51265, 'Самара': 50994, 'Пермь': 48089,
#                            'Нижний Новгород': 47662}
# dict_count_area = {'Москва': 0.3246, 'Санкт-Петербург': 0.1197, 'Новосибирск': 0.0271, 'Казань': 0.0237,
#                           'Нижний Новгород': 0.0232, 'Ростов-на-Дону': 0.0209, 'Екатеринбург': 0.0207,
#                           'Краснодар': 0.0185, 'Самара': 0.0143, 'Воронеж': 0.0141}

dict_salary = {2007: 38916, 2008: 43646, 2009: 42492, 2010: 43846, 2011: 47451, 2012: 48243, 2013: 51510, 2014: 50658}
dict_count = {2007: 2196, 2008: 17549, 2009: 17709, 2010: 29093, 2011: 36700, 2012: 44153, 2013: 59954, 2014: 66837}
dict_salary_vac = {2007: 43770, 2008: 50412, 2009: 46699, 2010: 50570, 2011: 55770, 2012: 57960, 2013: 58804, 2014: 62384}
dict_count_vac = {2007: 317, 2008: 2460, 2009: 2066, 2010: 3614, 2011: 4422, 2012: 4966, 2013: 5990, 2014: 5492}
dict_salary_area = {'Москва': 57354, 'Санкт-Петербург': 46291, 'Новосибирск': 41580, 'Екатеринбург': 41091, 'Казань': 37587, 'Самара': 34091, 'Нижний Новгород': 33637, 'Ярославль': 32744, 'Краснодар': 32542, 'Воронеж': 29725}
dict_count_area = {'Москва': 0.4581, 'Санкт-Петербург': 0.1415, 'Нижний Новгород': 0.0269, 'Казань': 0.0266, 'Ростов-на-Дону': 0.0234, 'Новосибирск': 0.0202, 'Екатеринбург': 0.0143, 'Воронеж': 0.014, 'Самара': 0.0133, 'Краснодар': 0.0131}



print("Динамика уровня зарплат по годам: " + str(dict_salary))
print("Динамика количества вакансий по годам: " + str(dict_count))
print("Динамика уровня зарплат по годам для выбранной профессии: " + str(dict_salary_vac))
print("Динамика количества вакансий по годам для выбранной профессии: " + str(dict_count_vac))
print("Уровень зарплат по городам (в порядке убывания): " + str(dict_salary_area))
print("Доля вакансий по городам (в порядке убывания): " + str(dict_count_area))
wb = Workbook()
del wb['Sheet']
sheet = wb.create_sheet('Статистика по годам')
thin = Side(border_style="thin", color="000000")
sheet["A1"] = "Год"
sheet["A1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
sheet["B1"] = "Средняя зарплата"
sheet["B1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
sheet["C1"] = "Средняя зарплата - Программист"
sheet["C1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
sheet["D1"] = "Количество вакансий"
sheet["D1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
sheet["E1"] = "Количество вакансий - Программист"
sheet["E1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
sheet["B1"].font = Font(bold=True)
sheet["C1"].font = Font(bold=True)
sheet["A1"].font = Font(bold=True)
sheet["D1"].font = Font(bold=True)
sheet["E1"].font = Font(bold=True)


def text_as(value):
    if value is None:
        return ""
    return str(value)



for row, (year, value) in enumerate(dict_salary.items(), start=2):
    sheet [f"A{row}"] = year
    sheet [f"A{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    sheet [f"B{row}"] = value
    sheet [f"B{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    sheet [f"C{row}"] = dict_salary_vac[year]
    sheet [f"C{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    sheet [f"D{row}"] = dict_count[year]
    sheet [f"D{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    sheet [f"E{row}"] = dict_count_vac[year]
    sheet [f"E{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)

for column_cells in sheet.columns:
    length = max(len(text_as(cell.value)) for cell in column_cells)
    sheet.column_dimensions[column_cells[0].column_letter].width = length+2

sheet = wb.create_sheet('Статистика по городам')
sheet["A1"] = "Город"
sheet["A1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
sheet["B1"] = "Уровень зарплат"
sheet["B1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
sheet["D1"] = "Город"
sheet["D1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
sheet["E1"] = "Доля вакансий"
sheet["E1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
sheet["B1"].font = Font(bold=True)
sheet["C1"].font = Font(bold=True)
sheet["A1"].font = Font(bold=True)
sheet["D1"].font = Font(bold=True)
sheet["E1"].font = Font(bold=True)
for row, (year, value) in enumerate(dict_salary_area.items(), start=2):
    sheet [f"A{row}"] = year
    sheet [f"A{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    sheet [f"B{row}"] = value
    sheet [f"B{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
for row, (year, value) in enumerate(dict_count_area.items(), start=2):
    sheet [f"D{row}"] = year
    sheet [f"D{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    sheet [f"E{row}"] = value
    sheet[f"E{row}"].number_format = FORMAT_PERCENTAGE_00
    sheet [f"E{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)

for column_cells in sheet.columns:
    length = max(len(text_as(cell.value)) for cell in column_cells)
    sheet.column_dimensions[column_cells[0].column_letter].width = length+2

wb.save('report.xlsx')